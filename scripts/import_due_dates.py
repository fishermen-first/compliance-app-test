#!/usr/bin/env python3
"""Generate Supabase SQL from the first tab of the client compliance workbook.

Usage:
  python scripts/import_due_dates.py \
    --workbook "Compliance Tracking.xlsx" \
    --company "Arctic Storm Management Group" \
    --out tmp/due-dates-import.sql

The workbook itself is ignored by git. This script only emits SQL and warnings.
"""

from __future__ import annotations

import argparse
import json
import re
from collections import Counter
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

HEADERS = [
    "owner_raw",
    "vessel",
    "item_name",
    "item_number",
    "agency_type",
    "frequency_label",
    "expiration_date",
    "start_working_on",
    "status_raw",
    "status_notes",
    "instructions",
]

STATUS_MAP = {
    "": "not_started",
    "in progress": "in_progress",
    "submitted": "submitted",
    "n/a": "discontinued",
    "na": "discontinued",
    "start soon": "not_started",
}

COMPANY_WIDE_NAMES = {"asmg", "ashco", "company", "office", ""}


def sql(value: Any) -> str:
    if value is None:
        return "null"
    if isinstance(value, (datetime, date)):
        value = value.date() if isinstance(value, datetime) else value
        return "'" + value.isoformat().replace("'", "''") + "'"
    text = str(value).strip()
    if not text or text.upper() == "NA":
        return "null"
    return "'" + text.replace("'", "''") + "'"


def clean(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value.date().isoformat() if isinstance(value, datetime) else value.isoformat()
    text = str(value).strip()
    return text or None


def date_value(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def import_date(value: Any, row_number: int, field_label: str, warnings: list[dict[str, Any]]) -> date | None:
    parsed = date_value(value)
    if parsed is None:
        warnings.append({"row": row_number, "issue": f"Missing or non-date {field_label}", "value": clean(value)})
        return None
    if parsed.year < 2000:
        warnings.append({"row": row_number, "issue": f"{field_label.title()} year looks like an outlier; set to null", "value": parsed.isoformat()})
        return None
    return parsed


def parse_owner_current(owner_raw: str | None) -> str | None:
    if not owner_raw:
        return None
    value = owner_raw.strip()
    if "-->" in value:
        return value.split("-->", 1)[0].strip()
    if "->" in value:
        return value.split("->", 1)[0].strip()
    if "/" in value:
        return value.split("/", 1)[0].strip()
    if "-" in value:
        return value.split("-", 1)[0].strip()
    return value


def infer_recurrence(frequency: str | None) -> tuple[str, int | None]:
    label = (frequency or "").strip().lower()
    if not label or label in {"na", "n/a"}:
        return "none", None
    if "unannounced" in label or "new permit" in label:
        return "manual", None
    if "quarter" in label:
        return "months", 3
    if "twice" in label:
        return "months", 6
    if "bienn" in label:
        return "years", 2
    if "trienn" in label:
        return "years", 3
    if "annual" in label:
        return "years", 1
    match = re.search(r"every\s+(\d+)\s+year", label)
    if match:
        return "years", int(match.group(1))
    match = re.search(r"every\s+(\d+)\s+month", label)
    if match:
        return "months", int(match.group(1))
    return "manual", None


def infer_area(agency_type: str | None, item_name: str | None) -> str:
    agency = (agency_type or "").lower()
    item = (item_name or "").lower()
    if "uscg" in agency or "fcc" in agency or "vessel" in item or "radio station" in item:
        return "Vessel Compliance"
    if any(x in agency for x in ["epa", "ecology", "chadux", "seapro"]) or any(x in item for x in ["oil", "discharge"]):
        return "Environmental"
    if any(x in agency for x in ["fda", "brcgs", "gfsi"]) or any(x in item for x in ["food", "haccp"]):
        return "Food Safety"
    if any(x in agency for x in ["msc", "rfm"]) or any(x in item for x in ["audit", "certificate"]):
        return "Audits & Certifications"
    if any(x in agency for x in ["noaa", "nmfs", "usdc"]):
        return "Fishing / Quota Reporting"
    if "drill" in item or "safety" in item:
        return "Safety / Drills"
    if "permit" in item or "license" in item:
        return "Permits & Licenses"
    return "Other"


def status_from_raw(value: str | None) -> str:
    return STATUS_MAP.get((value or "").strip().lower(), "not_started")


def load_records(workbook: Path, sheet_name: str | None) -> tuple[str, list[dict[str, Any]], list[dict[str, Any]]]:
    wb = load_workbook(workbook, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.worksheets[0]
    records: list[dict[str, Any]] = []
    warnings: list[dict[str, Any]] = []

    for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        values = list(row[: len(HEADERS)])
        if not any(cell is not None and str(cell).strip() != "" for cell in values):
            continue

        raw = dict(zip(HEADERS, values))
        item_name = clean(raw["item_name"])
        vessel = clean(raw["vessel"])
        if not item_name:
            warnings.append({"row": row_number, "issue": "Skipped row with no item name"})
            continue

        expiration = import_date(raw["expiration_date"], row_number, "expiration", warnings)
        start = import_date(raw["start_working_on"], row_number, "start working date", warnings)

        frequency = clean(raw["frequency_label"])
        recurrence_unit, recurrence_interval = infer_recurrence(frequency)
        owner_raw = clean(raw["owner_raw"])
        agency = clean(raw["agency_type"])

        records.append({
            "source_row_number": row_number,
            "owner_raw": owner_raw,
            "owner_current": parse_owner_current(owner_raw),
            "vessel": vessel,
            "item_name": item_name,
            "item_number": clean(raw["item_number"]),
            "agency_type": agency,
            "compliance_area": infer_area(agency, item_name),
            "frequency_label": frequency,
            "recurrence_unit": recurrence_unit,
            "recurrence_interval": recurrence_interval,
            "expiration_date": expiration,
            "start_working_on": start,
            "status": status_from_raw(clean(raw["status_raw"])),
            "status_notes": clean(raw["status_notes"]),
            "instructions": clean(raw["instructions"]),
        })

    return ws.title, records, warnings


def generate_sql(company_name: str, sheet_name: str, records: list[dict[str, Any]]) -> str:
    vessel_names = sorted({r["vessel"] for r in records if r["vessel"] and r["vessel"].strip().lower() not in COMPANY_WIDE_NAMES})
    lines = [
        "begin;",
        "",
        "do $$",
        "declare",
        "  target_company_id uuid;",
        "begin",
        f"  select id into target_company_id from public.companies where name = {sql(company_name)} order by created_at asc limit 1;",
        "",
        "  if target_company_id is null then",
        f"    insert into public.companies (name) values ({sql(company_name)}) returning id into target_company_id;",
        "  end if;",
        "",
    ]

    for vessel in vessel_names:
        lines.append(f"  insert into public.vessels (company_id, name) values (target_company_id, {sql(vessel)}) on conflict (company_id, name) do update set active = true, updated_at = now();")

    if vessel_names:
        lines.append("")

    for record in records:
        vessel = record["vessel"]
        vessel_expr = "null"
        if vessel and vessel.strip().lower() not in COMPANY_WIDE_NAMES:
            vessel_expr = f"(select id from public.vessels where company_id = target_company_id and name = {sql(vessel)} limit 1)"

        lines.append("  insert into public.compliance_items (")
        lines.append("    company_id, vessel_id, owner_raw, owner_current, item_name, item_number, agency_type, compliance_area,")
        lines.append("    frequency_label, recurrence_unit, recurrence_interval, start_working_on, expiration_date, status,")
        lines.append("    status_notes, instructions, source_sheet, source_row_number")
        lines.append("  ) values (")
        lines.append(f"    target_company_id, {vessel_expr}, {sql(record['owner_raw'])}, {sql(record['owner_current'])}, {sql(record['item_name'])}, {sql(record['item_number'])}, {sql(record['agency_type'])}, {sql(record['compliance_area'])},")
        lines.append(f"    {sql(record['frequency_label'])}, {sql(record['recurrence_unit'])}::public.recurrence_unit, {record['recurrence_interval'] if record['recurrence_interval'] is not None else 'null'}, {sql(record['start_working_on'])}, {sql(record['expiration_date'])}, {sql(record['status'])}::public.compliance_item_status,")
        lines.append(f"    {sql(record['status_notes'])}, {sql(record['instructions'])}, {sql(sheet_name)}, {record['source_row_number']}")
        lines.append("  )")
        lines.append("  on conflict (company_id, source_sheet, source_row_number)")
        lines.append("  where source_sheet is not null and source_row_number is not null")
        lines.append("  do update set")
        lines.append("    vessel_id = excluded.vessel_id,")
        lines.append("    owner_raw = excluded.owner_raw,")
        lines.append("    owner_current = excluded.owner_current,")
        lines.append("    item_name = excluded.item_name,")
        lines.append("    item_number = excluded.item_number,")
        lines.append("    agency_type = excluded.agency_type,")
        lines.append("    compliance_area = excluded.compliance_area,")
        lines.append("    frequency_label = excluded.frequency_label,")
        lines.append("    recurrence_unit = excluded.recurrence_unit,")
        lines.append("    recurrence_interval = excluded.recurrence_interval,")
        lines.append("    start_working_on = excluded.start_working_on,")
        lines.append("    expiration_date = excluded.expiration_date,")
        lines.append("    status_notes = excluded.status_notes,")
        lines.append("    instructions = excluded.instructions,")
        lines.append("    updated_at = now();")
        lines.append("")

    lines.extend([
        "end $$;",
        "",
        "insert into public.compliance_item_reminder_rules (item_id, company_id, label, trigger_type)",
        "select id, company_id, 'Start working reminder', 'on_start_date'",
        "from public.compliance_items",
        f"where source_sheet = {sql(sheet_name)}",
        "on conflict do nothing;",
        "",
        "insert into public.compliance_item_reminder_rules (item_id, company_id, label, trigger_type, days_before)",
        "select id, company_id, '14 days before expiration', 'days_before_expiration', 14",
        "from public.compliance_items",
        f"where source_sheet = {sql(sheet_name)}",
        "on conflict do nothing;",
        "",
        "commit;",
        "",
    ])
    return "\n".join(lines)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", default="Compliance Tracking.xlsx")
    parser.add_argument("--sheet", default=None)
    parser.add_argument("--company", required=True)
    parser.add_argument("--out", default="tmp/due-dates-import.sql")
    parser.add_argument("--warnings", default="tmp/due-dates-import-warnings.json")
    args = parser.parse_args()

    workbook = Path(args.workbook).resolve()
    sheet_name, records, warnings = load_records(workbook, args.sheet)

    out = Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(generate_sql(args.company, sheet_name, records), encoding="utf-8")

    warning_out = Path(args.warnings)
    warning_out.parent.mkdir(parents=True, exist_ok=True)
    summary = {
        "workbook": str(workbook),
        "sheet": sheet_name,
        "record_count": len(records),
        "vessels": Counter(r["vessel"] or "Company-wide" for r in records).most_common(),
        "statuses": Counter(r["status"] for r in records).most_common(),
        "compliance_areas": Counter(r["compliance_area"] for r in records).most_common(),
        "warnings": warnings,
    }
    warning_out.write_text(json.dumps(summary, indent=2, default=str), encoding="utf-8")

    print(f"Wrote {out} with {len(records)} records from {sheet_name}.")
    print(f"Wrote {warning_out} with {len(warnings)} warnings.")


if __name__ == "__main__":
    main()
